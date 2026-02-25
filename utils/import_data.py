import pandas as pd
from models import db
from models.planilha import PlanilhaData, AbaConfig
from models.user import User
import openpyxl

def import_excel_data(app, excel_file='MONITORAMENTO UECI - CONSOLIDADO.xlsx'):
    """Importa dados do arquivo Excel para o banco de dados"""
    
    with app.app_context():
        print("Iniciando importação de dados do Excel...")
        
        # Abas que devem ser ignoradas
        ABAS_IGNORADAS = [
            'Sugestões de iniciativa da área',
            'Dados lista suspensa',
            'T Dinâmica'
        ]
        
        # Carregar o arquivo Excel com openpyxl para ler cabeçalhos originais
        workbook = openpyxl.load_workbook(excel_file)
        
        print(f"Abas encontradas: {workbook.sheetnames}")
        
        # Processar cada aba
        for idx, sheet_name in enumerate(workbook.sheetnames):
            # Pular abas ignoradas
            if sheet_name in ABAS_IGNORADAS:
                print(f"\n⏭️  Ignorando aba: {sheet_name}")
                continue
            
            print(f"\n✓ Processando aba: {sheet_name}")
            
            worksheet = workbook[sheet_name]
            
            # Encontrar a linha do cabeçalho (procurar a primeira linha com dados)
            header_row_idx = None
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=20), start=1):
                # Contar células não vazias
                non_empty = sum(1 for cell in row if cell.value is not None and str(cell.value).strip())
                if non_empty > 3:  # Se tem mais de 3 células com conteúdo
                    header_row_idx = row_idx
                    break
            
            if header_row_idx is None:
                print(f"  ⚠️  Cabeçalho não encontrado na aba {sheet_name}, pulando...")
                continue
            
            # Ler cabeçalhos originais da linha identificada
            header_row = list(worksheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx))[0]
            original_headers = []
            
            for idx_col, cell in enumerate(header_row):
                if cell.value is not None and str(cell.value).strip():
                    original_headers.append(str(cell.value).strip())
                else:
                    original_headers.append(f'Coluna_{idx_col}')
            
            # Agora usar pandas para ler os dados, mas com o cabeçalho já identificado
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row_idx-1)
            
            # Aplicar os cabeçalhos originais
            if len(original_headers) <= len(df.columns):
                # Garantir que temos cabeçalhos para todas as colunas
                while len(original_headers) < len(df.columns):
                    original_headers.append(f'Coluna_{len(original_headers)}')
                df.columns = original_headers[:len(df.columns)]
            
            # Remover linhas completamente vazias
            df = df.dropna(how='all')
            
            print(f"  📋 Colunas identificadas: {df.columns.tolist()[:5]}{'...' if len(df.columns) > 5 else ''}")
            print(f"  📊 Total de linhas: {len(df)}")
            
            # Criar ou atualizar configuração da aba
            aba_config = AbaConfig.query.filter_by(aba_name=sheet_name).first()
            if not aba_config:
                aba_config = AbaConfig(
                    aba_name=sheet_name,
                    display_order=idx
                )
                db.session.add(aba_config)
            
            # Configurar colunas
            columns = []
            for col in df.columns:
                columns.append({
                    'name': col,
                    'type': 'text'
                })
            aba_config.set_columns(columns)
            
            # Importar dados (limitando a 100 linhas por aba para não sobrecarregar)
            max_rows = min(100, len(df))
            
            for row_idx, row in df.head(max_rows).iterrows():
                # Converter linha para dicionário
                row_data = {}
                for col in df.columns:
                    value = row[col]
                    if pd.notna(value):
                        row_data[col] = str(value)
                    else:
                        row_data[col] = ''
                
                # Criar registro
                planilha_data = PlanilhaData(
                    aba_name=sheet_name,
                    row_order=row_idx + 1,
                    created_by=1  # Admin
                )
                planilha_data.set_data(row_data)
                db.session.add(planilha_data)
            
            print(f"Importadas {max_rows} linhas da aba {sheet_name}")
        
        # Commit todas as mudanças
        db.session.commit()
        print("\n✓ Importação concluída com sucesso!")


def create_initial_users(app):
    """Cria os usuários iniciais do sistema"""
    
    with app.app_context():
        print("\nCriando usuários iniciais...")
        
        # Verificar se já existem usuários
        if User.query.count() > 0:
            print("Usuários já existem no banco de dados.")
            return
        
        # Criar administrador
        admin = User(
            username='admin',
            email='admin@ueci.es.gov.br',
            phone='27999999999',
            is_admin=True
        )
        admin.set_password('admin123')  # Senha temporária
        db.session.add(admin)
        
        # Criar usuárias
        users_data = [
            {'username': 'Larissa', 'email': 'larissa@ueci.es.gov.br', 'phone': '27988888888'},
            {'username': 'Carla', 'email': 'carla@ueci.es.gov.br', 'phone': '27977777777'},
            {'username': 'Gabriela', 'email': 'gabriela@ueci.es.gov.br', 'phone': '27966666666'}
        ]
        
        for user_data in users_data:
            user = User(
                username=user_data['username'],
                email=user_data['email'],
                phone=user_data['phone'],
                is_admin=False
            )
            # Usuários sem senha - precisarão definir via token
            db.session.add(user)
        
        db.session.commit()
        print("✓ Usuários criados com sucesso!")
        print("\nUsuário Admin criado:")
        print("  Username: admin")
        print("  Senha: admin123")
        print("\nUsuários criados (sem senha - use token para definir):")
        for user_data in users_data:
            print(f"  - {user_data['username']} ({user_data['email']})")
